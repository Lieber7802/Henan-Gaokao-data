from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

OPENING_BRACKETS = "（([【"
CLOSING_BRACKETS = "）)]】"
FORMULA_ERROR_MARKERS = ("#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#N/A")


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def bracket_balance(text: str) -> int:
    return sum(text.count(char) for char in OPENING_BRACKETS) - sum(text.count(char) for char in CLOSING_BRACKETS)


def is_implausible_numeric(batch: str, score: Any, rank: Any) -> bool:
    if not isinstance(score, (int, float)) or not isinstance(rank, (int, float)):
        return False
    if "本科一批" in batch:
        return score < 450 or (rank < 50 and score < 690)
    if "本科二批" in batch:
        return score < 300
    return score < 100 or rank < 1


def row_preview(row_index: int, values: list[Any]) -> dict[str, Any]:
    return {
        "row": row_index,
        "学校": values[0],
        "专业": values[1],
        "备注": values[2],
        "2024最低录取分数": values[4],
        "2024最低录取位次": values[5],
        "2023最低录取分数": values[7],
        "2023最低录取位次": values[8],
    }


def scan_workbook(path: Path, batch: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook.active
    checks: dict[str, list[dict[str, Any]]] = {
        "missing_school_marker": [],
        "unbalanced_major": [],
        "single_char_major": [],
        "bad_formula_prefix": [],
        "formula_error_text": [],
        "implausible_numeric": [],
    }
    rows_scanned = 0

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=10, values_only=True), start=2):
        rows_scanned += 1
        values = [value if value is not None else "" for value in row]
        school = clean_text(values[0])
        major = clean_text(values[1])

        if school and "大学" not in school and "学院" not in school:
            checks["missing_school_marker"].append(row_preview(row_index, values))

        if major and bracket_balance(major) != 0:
            checks["unbalanced_major"].append(row_preview(row_index, values))

        if len([char for char in major if "\u4e00" <= char <= "\u9fff"]) == 1:
            checks["single_char_major"].append(row_preview(row_index, values))

        for col_index, value in enumerate(values, start=1):
            if isinstance(value, str):
                if value.startswith("=="):
                    item = row_preview(row_index, values)
                    item["cell"] = f"{col_index}:{value}"
                    checks["bad_formula_prefix"].append(item)
                if any(marker in value for marker in FORMULA_ERROR_MARKERS):
                    item = row_preview(row_index, values)
                    item["cell"] = f"{col_index}:{value}"
                    checks["formula_error_text"].append(item)

        for year, score, rank in (("2024", values[4], values[5]), ("2023", values[7], values[8])):
            if is_implausible_numeric(batch, score, rank):
                item = row_preview(row_index, values)
                item["year"] = year
                checks["implausible_numeric"].append(item)

    return {
        "workbook": str(path),
        "batch": batch,
        "rows_scanned": rows_scanned,
        "issue_counts": {key: len(value) for key, value in checks.items()},
        "samples": {key: value[:20] for key, value in checks.items() if value},
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Check generated major-history workbook quality.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--batch", default="本科一批", help="Batch name, e.g. 本科一批 or 本科二批.")
    parser.add_argument("--output-json", type=Path, help="Optional path to write the scan result JSON.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = scan_workbook(args.workbook, args.batch)
    text = json.dumps(result, ensure_ascii=False, indent=2)
    print(text)
    if args.output_json:
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(text, encoding="utf-8")


if __name__ == "__main__":
    main()
