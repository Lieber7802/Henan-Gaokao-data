from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_COLUMNS = [
    "学校",
    "专业",
    "",
    "年份",
    "最低录取分数",
    "最低录取位次",
    "年份",
    "最低录取分数",
    "最低录取位次",
    "平均最低录取位次",
]


def load_payload(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def build_workbook(payload: dict[str, Any]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    columns = payload.get("columns") or DEFAULT_COLUMNS
    sheet.append(columns)
    for row_index, row in enumerate(payload.get("wide_rows") or [], start=2):
        sheet.append(
            [
                row.get("学校"),
                row.get("专业"),
                row.get("备注"),
                row.get("2024年份"),
                row.get("2024最低录取分数"),
                row.get("2024最低录取位次"),
                row.get("2023年份"),
                row.get("2023最低录取分数"),
                row.get("2023最低录取位次"),
                f'=IF(COUNT(F{row_index},I{row_index})=0,"",AVERAGE(F{row_index},I{row_index}))',
            ]
        )

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    border_side = Side(style="thin", color="D9E2EC")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in sheet[1]:
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="000000")
        cell.fill = header_fill
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, color="1F1F1F")
            cell.border = border

    widths = [22, 28, 24, 10, 14, 16, 10, 14, 16, 18]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    for column_index in [4, 5, 6, 7, 8, 9, 10]:
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for item in cell:
                item.number_format = "#,##0"

    sheet.freeze_panes = "A2"
    return workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build major-history Excel workbook from wide-table JSON.")
    parser.add_argument("data_json", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    payload = load_payload(args.data_json)
    workbook = build_workbook(payload)
    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output_xlsx)
    print(args.output_xlsx)


if __name__ == "__main__":
    main()
