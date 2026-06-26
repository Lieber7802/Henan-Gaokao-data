from __future__ import annotations

import argparse
import json
import re
import subprocess
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - reported in runtime validation.
    load_workbook = None

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "configs" / "ppstructure_pilot_sections.json"
OUTPUT_DIR = ROOT / "output" / "ppstructure_pilot"
RAW_DIR = OUTPUT_DIR / "raw"
PREVIEW_DIR = OUTPUT_DIR / "preview"
REVIEW_DIR = OUTPUT_DIR / "review"
TIMING_PATH = REVIEW_DIR / "ppstructure_timing.json"
PREVIEW_DATA_PATH = REVIEW_DIR / "ppstructure_preview_data.json"
REPORT_PATH = REVIEW_DIR / "ppstructure_decision_report.md"
WORKBOOK_BUILDER = ROOT / "workbook_build" / "build_ppstructure_final_preview.mjs"
WORKBOOK_PATH = PREVIEW_DIR / "ppstructure_final_preview.xlsx"

APPLICATION_COLUMNS = [
    "批次",
    "院校代码",
    "院校名称",
    "年份",
    "公布计划",
    "实际投档人数",
    "投档最低总分",
    "语文",
    "数学",
    "外语听力",
    "与分数线差值",
    "最低分位次",
    "source_page",
    "review_status",
]

MAJOR_COLUMNS = [
    "批次",
    "院校代码",
    "院校名称",
    "年份",
    "专业代码",
    "专业名称",
    "公布计划",
    "录取人数",
    "最高分",
    "平均分",
    "最低分",
    "与分数线差值",
    "最低分位次",
    "source_page",
    "review_status",
]

COLLEGE_COLUMNS = [
    "批次",
    "院校代码",
    "院校名称",
    "年份",
    "公布计划",
    "录取人数",
    "最高分",
    "平均分",
    "最低分",
    "平均分与分数线差值",
    "与分数线差值",
    "最低分位次",
    "source_page",
    "review_status",
]

SCORE_LINE_COLUMNS = ["类别", "年份", "批次", "分数线", "source_page", "review_status"]
SCORE_BAND_COLUMNS = ["分数", "人数", "累计人数", "年份", "科类", "source_page", "review_status"]
ISSUE_COLUMNS = ["PDF页", "章节", "问题类型", "说明", "建议动作", "source_page", "review_status"]
SUMMARY_COLUMNS = ["项目", "结论"]


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell_parts: list[str] | None = None
        self._cell_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell_parts = []
            self._cell_depth += 1
        elif tag == "br" and self._cell_parts is not None:
            self._cell_parts.append(" ")

    def handle_data(self, data: str) -> None:
        if self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._row is not None and self._cell_parts is not None:
            text = clean_text("".join(self._cell_parts))
            self._row.append(text)
            self._cell_parts = None
            self._cell_depth = max(0, self._cell_depth - 1)
        elif tag == "tr" and self._row is not None:
            if any(cell for cell in self._row):
                self.rows.append(self._row)
            self._row = None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def digits_only(value: str) -> str:
    return re.sub(r"\D+", "", value or "")


def load_config() -> dict[str, Any]:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def load_timing() -> list[dict[str, Any]]:
    if not TIMING_PATH.exists():
        return []
    return json.loads(TIMING_PATH.read_text(encoding="utf-8"))


def page_stem(page: int) -> str:
    return f"page_{page:03d}-{page:03d}"


def section_dirs(section_id: str) -> dict[str, Path]:
    return {
        "json": RAW_DIR / section_id / "json",
        "xlsx": RAW_DIR / section_id / "xlsx",
    }


def parse_html_table(html: str) -> list[list[str]]:
    parser = SimpleTableParser()
    parser.feed(html or "")
    return [row for row in parser.rows if any(cell for cell in row)]


def matrix_from_json(json_path: Path) -> list[list[list[str]]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    res = data.get("res", data)
    matrices: list[list[list[str]]] = []
    for table in res.get("table_res_list") or []:
        if not isinstance(table, dict):
            continue
        html = table.get("pred_html")
        if html:
            matrix = parse_html_table(str(html))
            if matrix:
                matrices.append(matrix)
    return matrices


def matrices_from_xlsx(xlsx_path: Path) -> list[list[list[str]]]:
    if load_workbook is None:
        return []
    matrices: list[list[list[str]]] = []
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    for worksheet in workbook.worksheets:
        matrix: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            cells = [clean_text(cell) for cell in row]
            while cells and not cells[-1]:
                cells.pop()
            if any(cells):
                matrix.append(cells)
        if matrix:
            matrices.append(matrix)
    workbook.close()
    return matrices


def table_ocr_items_from_json(json_path: Path) -> list[dict[str, Any]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    res = data.get("res", data)
    items: list[dict[str, Any]] = []
    for table_index, table in enumerate(res.get("table_res_list") or [], start=1):
        if not isinstance(table, dict):
            continue
        table_ocr = table.get("table_ocr_pred") or {}
        tokens = [clean_text(token) for token in table_ocr.get("rec_texts") or []]
        tokens = [token for token in tokens if token]
        if tokens:
            items.append({"source": "json_table_ocr", "path": str(json_path), "table_index": table_index, "tokens": tokens})
        html = table.get("pred_html")
        if html:
            matrix = parse_html_table(str(html))
            if matrix:
                items.append({"source": "json_pred_html", "path": str(json_path), "table_index": table_index, "matrix": matrix})

    if not items:
        overall = res.get("overall_ocr_res") or {}
        tokens = [clean_text(token) for token in overall.get("rec_texts") or []] if isinstance(overall, dict) else []
        tokens = [token for token in tokens if token]
        if tokens:
            items.append({"source": "json_overall_ocr", "path": str(json_path), "tokens": tokens})
    return items


def load_page_items(section_id: str, page: int) -> tuple[list[dict[str, Any]], list[str]]:
    dirs = section_dirs(section_id)
    stem = page_stem(page)
    xlsx_items: list[dict[str, Any]] = []
    json_items: list[dict[str, Any]] = []
    issues: list[str] = []

    json_files = sorted(dirs["json"].glob(f"{stem}*.json"))
    for json_path in json_files:
        try:
            json_items.extend(table_ocr_items_from_json(json_path))
        except Exception as exc:  # noqa: BLE001
            issues.append(f"解析 JSON 失败：{json_path.name}: {type(exc).__name__}: {exc}")

    xlsx_files = sorted(dirs["xlsx"].glob(f"{stem}*.xlsx"))
    for xlsx_path in xlsx_files:
        try:
            for matrix in matrices_from_xlsx(xlsx_path):
                xlsx_items.append({"source": "xlsx", "path": str(xlsx_path), "matrix": matrix})
        except Exception as exc:  # noqa: BLE001
            issues.append(f"读取 XLSX 失败：{xlsx_path.name}: {type(exc).__name__}: {exc}")

    if not json_files:
        issues.append("缺少 PP-StructureV3 JSON 原始结果")
    if json_items:
        token_items = [item for item in json_items if "tokens" in item]
        if token_items:
            return token_items, issues
        return json_items, issues
    if xlsx_items:
        return xlsx_items, issues
    if not json_items and not xlsx_items:
        issues.append("未能从 XLSX 或 JSON pred_html 得到表格矩阵")
    return [], issues


def batch_label(section: dict[str, Any]) -> str:
    name = section["section_name_cn"]
    if "本科一批" in name:
        return "本科一批"
    if "本科二批" in name:
        return "本科二批"
    if "高职高专" in name:
        return "高职高专"
    return name


def likely_header_or_title(cells: list[str]) -> bool:
    text = "".join(cells)
    header_words = ["院校代号", "院校代码", "院校名称", "专业名称", "录取人数", "公布计划", "最低分", "分数段", "累计"]
    if any(word in text for word in header_words):
        digit_count = len(re.findall(r"\d", text))
        return digit_count < 12
    return False


def extract_year(cells: list[str], fallback: str = "") -> str:
    text = " ".join(cells)
    match = re.search(r"20(?:23|24|25)", text)
    return match.group(0) if match else fallback


def extract_school_code(cells: list[str]) -> tuple[str, int]:
    for index, cell in enumerate(cells[:5]):
        match = re.search(r"(?<!\d)(\d{4})(?!\d)", cell)
        if match:
            return match.group(1), index
    text = " ".join(cells)
    match = re.search(r"(?<!\d)(\d{4})(?!\d)", text)
    return (match.group(1), -1) if match else ("", -1)


def extract_school_name(cells: list[str], code_index: int) -> str:
    candidates = cells[code_index + 1 : code_index + 4] if code_index >= 0 else cells[:4]
    for cell in candidates:
        if re.search(r"[\u4e00-\u9fff]", cell) and not likely_header_or_title([cell]):
            return cell
    for cell in cells:
        if re.search(r"[\u4e00-\u9fff]", cell) and "年" not in cell and "批" not in cell:
            return cell
    return ""


def numeric_tokens(cells: list[str]) -> list[str]:
    tokens: list[str] = []
    for cell in cells:
        tokens.extend(re.findall(r"-?\d+(?:\.\d+)?", cell))
    return tokens


def first_profession_code(cells: list[str]) -> tuple[str, int]:
    for index, cell in enumerate(cells[:6]):
        stripped = cell.strip()
        if re.fullmatch(r"[A-Z0-9]{1,4}", stripped) and not re.fullmatch(r"20\d{2}", stripped) and not re.fullmatch(r"\d{4}", stripped):
            return stripped, index
        match = re.search(r"(?<!\w)([A-Z]?\d{1,3})(?!\w)", stripped)
        if match and len(match.group(1)) <= 3 and not re.fullmatch(r"20\d{2}", match.group(1)):
            return match.group(1), index
    return "", -1


def has_chinese_name(value: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]{2,}", value or ""))


def map_numbers(target: dict[str, Any], fields: list[str], nums: list[str]) -> None:
    for field, value in zip(fields, nums):
        target[field] = value


def is_year_token(token: str) -> bool:
    return bool(re.fullmatch(r"20(?:23|24|25)", token or ""))


def is_school_code_token(token: str) -> bool:
    return bool(re.fullmatch(r"\d{4}", token or "")) and not is_year_token(token)


def is_school_name_token(token: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", token or "")) and any(word in token for word in ["大学", "学院", "学校", "院"])


def numeric_token_values(tokens: list[str]) -> list[str]:
    values: list[str] = []
    for token in tokens:
        if re.fullmatch(r"-?\d+(?:\.\d+)?", token):
            values.append(token)
    return values


def app_row_from_segment(section: dict[str, Any], page: int, source: str, code: str, name: str, segment: list[str]) -> dict[str, Any] | None:
    if not segment or not is_year_token(segment[0]):
        return None
    nums = numeric_token_values(segment[1:])
    if len(nums) < 3:
        return None
    row = {column: "" for column in APPLICATION_COLUMNS}
    row.update(
        {
            "批次": batch_label(section),
            "院校代码": code,
            "院校名称": name,
            "年份": segment[0],
            "source_page": page,
        }
    )
    map_numbers(row, ["公布计划", "实际投档人数", "投档最低总分", "语文", "数学", "外语听力", "与分数线差值", "最低分位次"], nums[:8])
    row["review_status"] = f"ppstructure_{source}_candidate" if row["投档最低总分"] else f"needs_review_ppstructure_{source}"
    return row


def application_rows_from_tokens(section: dict[str, Any], tokens: list[str], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for index in range(len(tokens) - 1):
        code = tokens[index]
        name = tokens[index + 1]
        if not (is_school_code_token(code) and has_chinese_name(name)):
            continue

        backward_start = -1
        for probe in range(index - 1, max(-1, index - 14), -1):
            if is_year_token(tokens[probe]) and len(numeric_token_values(tokens[probe + 1 : index])) >= 8:
                backward_start = probe
                break
        if backward_start >= 0:
            row = app_row_from_segment(section, page, source, code, name, tokens[backward_start:index])
            if row:
                key = (row["院校代码"], row["院校名称"], row["年份"])
                if key not in seen:
                    rows.append(row)
                    seen.add(key)

        forward_start = index + 2
        if forward_start < len(tokens) and is_year_token(tokens[forward_start]):
            forward_end = min(len(tokens), forward_start + 12)
            for probe in range(forward_start + 1, min(len(tokens) - 1, forward_start + 13)):
                if is_school_code_token(tokens[probe]) and has_chinese_name(tokens[probe + 1]):
                    forward_end = probe
                    break
            row = app_row_from_segment(section, page, source, code, name, tokens[forward_start:forward_end])
            if row:
                key = (row["院校代码"], row["院校名称"], row["年份"])
                if key not in seen:
                    rows.append(row)
                    seen.add(key)
    return rows


def major_rows_from_tokens(section: dict[str, Any], tokens: list[str], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_school_code = ""
    current_school_name = ""
    current_year = ""
    index = 0
    while index < len(tokens):
        token = tokens[index]
        next_token = tokens[index + 1] if index + 1 < len(tokens) else ""
        if is_school_code_token(token) and has_chinese_name(next_token):
            current_school_code = token
            current_school_name = next_token
            index += 2
            continue
        if is_year_token(token):
            current_year = token
            index += 1
            continue
        if current_school_code and re.fullmatch(r"[A-Z0-9]{1,4}", token or "") and not is_school_code_token(token):
            prof_name = next_token if has_chinese_name(next_token) else ""
            if prof_name:
                lookahead = tokens[index + 2 : index + 14]
                nums = numeric_token_values([value for value in lookahead if not is_year_token(value)])
                if len(nums) >= 3:
                    row = {column: "" for column in MAJOR_COLUMNS}
                    row.update(
                        {
                            "批次": batch_label(section),
                            "院校代码": current_school_code,
                            "院校名称": current_school_name,
                            "年份": current_year,
                            "专业代码": token,
                            "专业名称": prof_name,
                            "source_page": page,
                            "review_status": f"needs_cross_row_postprocess_ppstructure_{source}",
                        }
                    )
                    map_numbers(row, ["公布计划", "录取人数", "最高分", "平均分", "最低分", "与分数线差值", "最低分位次"], nums[:7])
                    rows.append(row)
                    index += 2
                    continue
        index += 1
    first_year_by_school: dict[tuple[str, str], str] = {}
    for row in rows:
        key = (row["院校代码"], row["院校名称"])
        if row["年份"] and key not in first_year_by_school:
            first_year_by_school[key] = row["年份"]
    for row in rows:
        if not row["年份"]:
            row["年份"] = first_year_by_school.get((row["院校代码"], row["院校名称"]), "")
    return rows


def college_rows_from_tokens(section: dict[str, Any], tokens: list[str], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index in range(len(tokens) - 1):
        code = tokens[index]
        name = tokens[index + 1]
        if not (is_school_code_token(code) and has_chinese_name(name)):
            continue
        cursor = index + 2
        while cursor < len(tokens):
            if cursor + 1 < len(tokens) and is_school_code_token(tokens[cursor]) and has_chinese_name(tokens[cursor + 1]):
                break
            if not is_year_token(tokens[cursor]):
                cursor += 1
                continue
            next_break = min(len(tokens), cursor + 12)
            for probe in range(cursor + 1, len(tokens) - 1):
                if is_year_token(tokens[probe]) or (is_school_code_token(tokens[probe]) and has_chinese_name(tokens[probe + 1])):
                    next_break = probe
                    break
            nums = numeric_token_values(tokens[cursor + 1 : next_break])
            if len(nums) >= 6:
                row = {column: "" for column in COLLEGE_COLUMNS}
                row.update({"批次": batch_label(section), "院校代码": code, "院校名称": name, "年份": tokens[cursor], "source_page": page})
                mapping = {
                    "公布计划": nums[0],
                    "录取人数": nums[1],
                    "最高分": nums[2],
                    "最低分": nums[3],
                    "与分数线差值": nums[4] if len(nums) > 4 else "",
                    "最低分位次": nums[5] if len(nums) > 5 else "",
                    "平均分": nums[6] if len(nums) > 6 else "",
                    "平均分与分数线差值": nums[7] if len(nums) > 7 else "",
                }
                row.update(mapping)
                row["review_status"] = f"ppstructure_{source}_candidate"
                rows.append(row)
            cursor = next_break
    return rows


def score_rows_from_tokens(section: dict[str, Any], tokens: list[str], page: int, source: str) -> list[dict[str, Any]]:
    family = section["table_family"]
    if family == "score_lines":
        rows: list[dict[str, Any]] = []
        batch_names = ["本科一批", "本科二批", "高职高专批"]
        for index, token in enumerate(tokens):
            if token not in {"文科综合", "理科综合"}:
                continue
            nums = numeric_token_values(tokens[index + 1 : index + 4])
            if len(nums) < 3:
                continue
            year = ""
            for probe in range(index - 1, max(-1, index - 6), -1):
                if is_year_token(tokens[probe]):
                    year = tokens[probe]
                    break
            if not year:
                for probe in range(index + 4, min(len(tokens), index + 9)):
                    if is_year_token(tokens[probe]):
                        year = tokens[probe]
                        break
            for batch, score in zip(batch_names, nums):
                rows.append(
                    {
                        "类别": token,
                        "年份": year,
                        "批次": batch,
                        "分数线": score,
                        "source_page": page,
                        "review_status": f"ppstructure_{source}_candidate",
                    }
                )
        return rows
    rows = []
    for index, token in enumerate(tokens):
        if re.fullmatch(r"\d{3}", token or "") and index + 2 < len(tokens):
            nums = numeric_token_values(tokens[index : index + 5])
            if len(nums) >= 2:
                rows.append(
                    {
                        "分数": nums[0],
                        "人数": nums[1] if len(nums) > 1 else "",
                        "累计人数": nums[2] if len(nums) > 2 else "",
                        "年份": extract_year(tokens),
                        "科类": "理科",
                        "source_page": page,
                        "review_status": f"needs_review_ppstructure_{source}",
                    }
                )
    return rows


def application_rows(section: dict[str, Any], matrix: list[list[str]], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_school_code = ""
    current_school_name = ""
    current_year = ""
    for raw_cells in matrix:
        cells = [clean_text(cell) for cell in raw_cells if clean_text(cell)]
        if not cells or likely_header_or_title(cells):
            continue
        code, code_index = extract_school_code(cells)
        year = extract_year(cells, current_year)
        if code:
            current_school_code = code
            current_school_name = extract_school_name(cells, code_index) or current_school_name
        school_name = current_school_name or extract_school_name(cells, code_index)
        current_year = year or current_year
        nums = [value for value in numeric_tokens(cells) if value not in {code, year}]
        row = {column: "" for column in APPLICATION_COLUMNS}
        row.update(
            {
                "批次": batch_label(section),
                "院校代码": current_school_code,
                "院校名称": school_name,
                "年份": year,
                "source_page": page,
            }
        )
        map_numbers(row, ["公布计划", "实际投档人数", "投档最低总分", "语文", "数学", "外语听力", "与分数线差值", "最低分位次"], nums[:8])
        if row["院校代码"] and row["院校名称"] and row["投档最低总分"]:
            row["review_status"] = f"ppstructure_{source}_candidate"
        elif row["院校代码"] or row["院校名称"] or len(nums) >= 4:
            row["review_status"] = f"needs_review_ppstructure_{source}"
        else:
            continue
        rows.append(row)
    return rows


def major_rows(section: dict[str, Any], matrix: list[list[str]], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_school_code = ""
    current_school_name = ""
    current_year = ""
    for raw_cells in matrix:
        cells = [clean_text(cell) for cell in raw_cells if clean_text(cell)]
        if not cells or likely_header_or_title(cells):
            continue
        school_code, code_index = extract_school_code(cells)
        if school_code:
            current_school_code = school_code
            current_school_name = extract_school_name(cells, code_index) or current_school_name
        year = extract_year(cells, current_year)
        if year:
            current_year = year

        prof_code, prof_index = first_profession_code(cells)
        prof_name = ""
        for cell in cells[prof_index + 1 :] if prof_index >= 0 else cells:
            if has_chinese_name(cell) and not re.search(r"大学|学院|学校", cell):
                prof_name = cell
                break
        nums = [value for value in numeric_tokens(cells) if value not in {current_school_code, current_year, prof_code}]
        row = {column: "" for column in MAJOR_COLUMNS}
        row.update(
            {
                "批次": batch_label(section),
                "院校代码": current_school_code,
                "院校名称": current_school_name,
                "年份": current_year,
                "专业代码": prof_code,
                "专业名称": prof_name,
                "source_page": page,
            }
        )
        map_numbers(row, ["公布计划", "录取人数", "最高分", "平均分", "最低分", "与分数线差值", "最低分位次"], nums[:7])
        if row["专业代码"] and row["专业名称"] and row["最低分"]:
            row["review_status"] = f"ppstructure_{source}_candidate"
            rows.append(row)
        elif (row["专业代码"] or row["专业名称"] or len(nums) >= 4) and (current_school_code or current_school_name):
            row["review_status"] = f"needs_cross_row_postprocess_ppstructure_{source}"
            rows.append(row)
    return rows


def college_rows(section: dict[str, Any], matrix: list[list[str]], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw_cells in matrix:
        cells = [clean_text(cell) for cell in raw_cells if clean_text(cell)]
        if not cells or likely_header_or_title(cells):
            continue
        code, code_index = extract_school_code(cells)
        name = extract_school_name(cells, code_index)
        year = extract_year(cells)
        nums = [value for value in numeric_tokens(cells) if value not in {code, year}]
        row = {column: "" for column in COLLEGE_COLUMNS}
        row.update(
            {
                "批次": batch_label(section),
                "院校代码": code,
                "院校名称": name,
                "年份": year,
                "source_page": page,
            }
        )
        map_numbers(row, ["公布计划", "录取人数", "最高分", "平均分", "最低分", "平均分与分数线差值", "与分数线差值", "最低分位次"], nums[:8])
        if row["院校代码"] and row["院校名称"] and (row["最低分"] or row["最高分"]):
            row["review_status"] = f"ppstructure_{source}_candidate"
            rows.append(row)
        elif row["院校代码"] or row["院校名称"] or len(nums) >= 4:
            row["review_status"] = f"needs_review_ppstructure_{source}"
            rows.append(row)
    return rows


def score_line_rows(section: dict[str, Any], matrix: list[list[str]], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw_cells in matrix:
        cells = [clean_text(cell) for cell in raw_cells if clean_text(cell)]
        if not cells:
            continue
        text = " ".join(cells)
        nums = numeric_tokens(cells)
        if not nums:
            continue
        year = extract_year(cells)
        category = "理科" if "理" in text else ("文科" if "文" in text else "")
        for batch in ["本科一批", "本科二批", "高职高专"]:
            if batch in text:
                rows.append(
                    {
                        "类别": category,
                        "年份": year,
                        "批次": batch,
                        "分数线": nums[-1],
                        "source_page": page,
                        "review_status": f"needs_review_ppstructure_{source}",
                    }
                )
    return rows


def score_band_rows(section: dict[str, Any], matrix: list[list[str]], page: int, source: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw_cells in matrix:
        cells = [clean_text(cell) for cell in raw_cells if clean_text(cell)]
        if not cells or likely_header_or_title(cells):
            continue
        nums = numeric_tokens(cells)
        if len(nums) < 2:
            continue
        row = {
            "分数": nums[0],
            "人数": nums[1] if len(nums) > 1 else "",
            "累计人数": nums[2] if len(nums) > 2 else "",
            "年份": extract_year(cells),
            "科类": "理科",
            "source_page": page,
            "review_status": f"needs_review_ppstructure_{source}",
        }
        rows.append(row)
    return rows


def convert_matrix(section: dict[str, Any], matrix: list[list[str]], page: int, source: str) -> list[dict[str, Any]]:
    family = section["table_family"]
    if family == "application":
        return application_rows(section, matrix, page, source)
    if family == "major_admission":
        return major_rows(section, matrix, page, source)
    if family == "college_admission":
        return college_rows(section, matrix, page, source)
    if family == "score_lines":
        return score_line_rows(section, matrix, page, source)
    if family == "score_bands":
        return score_band_rows(section, matrix, page, source)
    return []


def convert_item(section: dict[str, Any], item: dict[str, Any], page: int) -> list[dict[str, Any]]:
    source = item["source"]
    if "tokens" in item:
        tokens = [clean_text(token) for token in item["tokens"] if clean_text(token)]
        family = section["table_family"]
        if family == "application":
            return application_rows_from_tokens(section, tokens, page, source)
        if family == "major_admission":
            return major_rows_from_tokens(section, tokens, page, source)
        if family == "college_admission":
            return college_rows_from_tokens(section, tokens, page, source)
        if family in {"score_lines", "score_bands"}:
            return score_rows_from_tokens(section, tokens, page, source)
        return []
    return convert_matrix(section, item["matrix"], page, source)


def columns_for_family(family: str) -> list[str]:
    if family == "application":
        return APPLICATION_COLUMNS
    if family == "major_admission":
        return MAJOR_COLUMNS
    if family == "college_admission":
        return COLLEGE_COLUMNS
    if family == "score_lines":
        return SCORE_LINE_COLUMNS
    if family == "score_bands":
        return SCORE_BAND_COLUMNS
    return ["source_page", "review_status"]


def make_issue(page: int, section: dict[str, Any], issue_type: str, detail: str, action: str, status: str = "needs_review") -> dict[str, Any]:
    return {
        "PDF页": page,
        "章节": section["section_name_cn"],
        "问题类型": issue_type,
        "说明": detail,
        "建议动作": action,
        "source_page": page,
        "review_status": status,
    }


def text_search(rows_by_sheet: dict[str, list[dict[str, Any]]], needle: str) -> bool:
    return needle in json.dumps(rows_by_sheet, ensure_ascii=False)


def build_summary(
    timing: list[dict[str, Any]],
    rows_by_sheet: dict[str, list[dict[str, Any]]],
    issues: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    ok_pages = [row for row in timing if row.get("status") == "ok"]
    pages_with_tables = [row for row in ok_pages if int(row.get("table_count") or 0) > 0 or row.get("has_xlsx") or row.get("has_html")]
    chapter2_rows = sum(len(rows_by_sheet.get(sheet, [])) for sheet in ["03_本科一批投档", "04_本科二批投档", "05_高职高专投档"])
    chapter3_rows = sum(len(rows_by_sheet.get(sheet, [])) for sheet in ["06_本科一批分专业录取", "07_本科二批分专业录取", "08_高职高专录取"])
    page44_key = text_search(rows_by_sheet, "首都体育学院") or text_search(rows_by_sheet, "天津体育学院")
    page123_key = text_search(rows_by_sheet, "北京体育大学")
    if chapter2_rows and page44_key:
        chapter2_decision = "第二章可进入全量 PP-StructureV3 路线，但仍需字段映射和抽查。"
    elif chapter2_rows:
        chapter2_decision = "第二章已有可读表格行，但页44关键院校未稳定进入最终字段，需要先修正后处理。"
    else:
        chapter2_decision = "第二章暂不能全量跑；当前缺少可读的院校级最终行。"
    if chapter3_rows:
        chapter3_decision = "第三章保留 PP-StructureV3 作为识别层，并增加院校/年份跨行继承后处理。"
    else:
        chapter3_decision = "第三章暂不能全量跑；需要确认 PP-StructureV3 是否识别出了分专业表格。"
    if pages_with_tables and chapter2_rows:
        custom_cut = "暂不需要回到自定义切格作为主路线；仅在个别失败页兜底。"
    else:
        custom_cut = "若补跑后仍无法得到可用表格结构，再考虑自定义切格兜底。"
    full_run = "建议先全量跑第二章；第三章做跨行继承后再试点扩页。" if chapter2_rows else "暂不建议全量跑，先修复 PP-StructureV3 输出或后处理解析。"
    return [
        {"项目": "PP-StructureV3 是否能识别出表格", "结论": f"试点成功页 {len(ok_pages)} 页，其中 {len(pages_with_tables)} 页有 table_res_list、HTML 或 XLSX 证据。"},
        {"项目": "直接导出的 XLSX 是否接近最终结构", "结论": "能稳定生成，但合并单元格和读序会错位，不能直接作为最终统计表；最终 Excel 需要基于 PP-StructureV3 输出做业务后处理。"},
        {"项目": "哪些章节可以进入全量跑", "结论": chapter2_decision},
        {"项目": "哪些章节需要后处理", "结论": f"第一章分数段统计需要专门分数段规则；{chapter3_decision}"},
        {"项目": "是否仍需要自定义切格方案", "结论": custom_cut},
        {"项目": "是否建议全量跑", "结论": full_run},
        {"项目": "页44业务抽查", "结论": "通过：找到首都体育学院/天津体育学院。" if page44_key else "未通过：最终预览字段中未找到首都体育学院或天津体育学院。"},
        {"项目": "页123业务抽查", "结论": "通过：找到北京体育大学。" if page123_key else "未通过：最终预览字段中未找到北京体育大学。"},
        {"项目": "问题页数量", "结论": str(len(issues))},
    ]


def build_decision_report(summary_rows: list[dict[str, Any]], issues: list[dict[str, Any]], timing: list[dict[str, Any]]) -> str:
    summary = {row["项目"]: row["结论"] for row in summary_rows}
    lines = [
        "# PP-StructureV3 章节级表格提取试点决策报告",
        "",
        "## 直接结论",
        "",
        f"- PP-StructureV3 是否能识别出表格：{summary.get('PP-StructureV3 是否能识别出表格', '')}",
        f"- 直接导出的 XLSX 是否接近最终结构：{summary.get('直接导出的 XLSX 是否接近最终结构', '')}",
        f"- 哪些章节可以进入全量跑：{summary.get('哪些章节可以进入全量跑', '')}",
        f"- 哪些章节需要后处理：{summary.get('哪些章节需要后处理', '')}",
        f"- 是否仍需要自定义切格方案：{summary.get('是否仍需要自定义切格方案', '')}",
        f"- 是否建议全量跑：{summary.get('是否建议全量跑', '')}",
        "",
        "## 运行证据",
        "",
        "| PDF页 | 章节 | 状态 | 表格数 | XLSX | HTML | JSON | 用时秒 |",
        "|---:|---|---|---:|---|---|---|---:|",
    ]
    for row in sorted(timing, key=lambda item: int(item.get("pdf_page", 0))):
        lines.append(
            "| {page} | {section} | {status} | {tables} | {xlsx} | {html} | {json} | {seconds} |".format(
                page=row.get("pdf_page", ""),
                section=row.get("section_name_cn", row.get("section_id", "")),
                status=row.get("status", ""),
                tables=row.get("table_count", ""),
                xlsx="Y" if row.get("has_xlsx") else "",
                html="Y" if row.get("has_html") else "",
                json="Y" if row.get("has_json") else "",
                seconds=row.get("elapsed_seconds", ""),
            )
        )
    lines.extend(["", "## 问题页清单", ""])
    if not issues:
        lines.append("暂无问题页。")
    else:
        lines.extend(["| PDF页 | 章节 | 问题类型 | 说明 | 建议动作 |", "|---:|---|---|---|---|"])
        for issue in issues:
            lines.append(
                f"| {issue['PDF页']} | {issue['章节']} | {issue['问题类型']} | {issue['说明']} | {issue['建议动作']} |"
            )
    return "\n".join(lines) + "\n"


def build_preview_data() -> dict[str, Any]:
    config = load_config()
    timing = load_timing()
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
    columns_by_sheet: dict[str, list[str]] = {}
    issues: list[dict[str, Any]] = []

    for section in config["sections"]:
        sheet = section["expected_final_sheet"]
        columns_by_sheet[sheet] = columns_for_family(section["table_family"])
        rows_by_sheet.setdefault(sheet, [])
        for page in section["pdf_pages"]:
            items, page_issues = load_page_items(section["section_id"], int(page))
            for message in page_issues:
                issues.append(make_issue(int(page), section, "表格提取", message, "检查 PP-StructureV3 原始 JSON/HTML/XLSX。"))
            for item in items:
                converted = convert_item(section, item, int(page))
                if converted:
                    rows_by_sheet[sheet].extend(converted)
                else:
                    issues.append(
                        make_issue(
                            int(page),
                            section,
                            "字段后处理",
                            f"PP-StructureV3 生成了表格矩阵，但未能稳定映射到 {sheet} 的最终业务字段。",
                            "保留原始 XLSX/HTML，调整该表类字段映射规则。",
                        )
                    )

    postprocess_issue_keys: set[tuple[str, int, str]] = set()
    for section in config["sections"]:
        sheet = section["expected_final_sheet"]
        for row in rows_by_sheet.get(sheet, []):
            status = str(row.get("review_status", ""))
            if not status.startswith("needs_"):
                continue
            page = int(row.get("source_page") or 0)
            key = (sheet, page, status)
            if key in postprocess_issue_keys:
                continue
            postprocess_issue_keys.add(key)
            issue_type = "字段后处理"
            detail = f"{sheet} 第 {page} 页存在 {status} 记录，需要规则整理后才能作为最终统计数据。"
            action = "保留 PP-StructureV3 识别结果，补充字段映射、跨行继承或分数段专用解析。"
            section_for_page = next((item for item in config["sections"] if page in item["pdf_pages"]), section)
            issues.append(make_issue(page, section_for_page, issue_type, detail, action, status))

    summary_rows = build_summary(timing, rows_by_sheet, issues)
    rows_by_sheet["00_试点结论"] = summary_rows
    columns_by_sheet["00_试点结论"] = SUMMARY_COLUMNS
    rows_by_sheet["09_问题页清单"] = issues
    columns_by_sheet["09_问题页清单"] = ISSUE_COLUMNS

    sheet_order = [
        "00_试点结论",
        "01_录取控制分数线",
        "02_分数段统计",
        "03_本科一批投档",
        "04_本科二批投档",
        "05_高职高专投档",
        "06_本科一批分专业录取",
        "07_本科二批分专业录取",
        "08_高职高专录取",
        "09_问题页清单",
    ]
    data = {
        "sheet_order": sheet_order,
        "columns_by_sheet": columns_by_sheet,
        "rows_by_sheet": rows_by_sheet,
        "timing": timing,
    }
    return data


def write_outputs(data: dict[str, Any], build_workbook: bool) -> None:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    PREVIEW_DATA_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    report = build_decision_report(data["rows_by_sheet"]["00_试点结论"], data["rows_by_sheet"]["09_问题页清单"], data["timing"])
    REPORT_PATH.write_text(report, encoding="utf-8")
    if build_workbook:
        result = subprocess.run(["node", str(WORKBOOK_BUILDER)], cwd=str(ROOT), check=False)
        if result.returncode != 0 and not WORKBOOK_PATH.exists():
            raise subprocess.CalledProcessError(result.returncode, result.args)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build final-format PP-StructureV3 pilot preview workbook.")
    parser.add_argument("--no-workbook", action="store_true", help="Only write preview_data.json and markdown report.")
    return parser


def main() -> None:
    args = build_arg_parser().parse_args()
    data = build_preview_data()
    write_outputs(data, build_workbook=not args.no_workbook)
    print(WORKBOOK_PATH)
    print(REPORT_PATH)


if __name__ == "__main__":
    main()
