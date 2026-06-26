import json
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "00_试点结论",
    "01_录取控制分数线",
    "02_分数段统计",
    "03_本科一批投档",
    "04_本科二批投档",
    "05_高职高专投档",
    "06_本科一批分专业录取",
    "07_本科二批分专业录取",
    "08_高职高专录取",
    "09_问题页清单",
]


def test_ppstructure_section_config_groups_23_pages():
    config = json.loads(Path("configs/ppstructure_pilot_sections.json").read_text(encoding="utf-8"))
    sections = config["sections"]
    pages = [page for section in sections for page in section["pdf_pages"]]
    assert len(sections) == 8
    assert len(pages) == 23
    assert len(set(pages)) == 23
    assert sorted(pages) == [5, 6, 10, 18, 20, 21, 22, 43, 44, 45, 82, 83, 122, 123, 124, 125, 456, 457, 458, 869, 870, 871, 912]
    for section in sections:
        assert {"section_id", "section_name_cn", "pdf_pages", "expected_final_sheet", "table_family"} <= set(section)


def test_ppstructure_preview_workbook_contract_after_run():
    workbook_path = Path("output/ppstructure_pilot/preview/ppstructure_final_preview.xlsx")
    if not workbook_path.exists():
        return
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == EXPECTED_SHEETS
        all_text = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            all_text.extend(str(cell) for row in rows for cell in row if cell is not None)
            if sheet_name not in {"00_试点结论", "09_问题页清单"}:
                headers = [cell for cell in rows[0] if cell]
                assert "source_page" in headers
                assert "review_status" in headers
        joined = "\n".join(all_text)
        assert "bbox_json" not in joined
        assert "poly_json" not in joined
        assert "ocr_index" not in joined
        conclusion_rows = list(workbook["00_试点结论"].iter_rows(values_only=True))
        assert any("是否建议全量跑" in str(cell) for row in conclusion_rows for cell in row if cell)
    finally:
        workbook.close()
